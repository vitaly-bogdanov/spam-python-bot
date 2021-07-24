import os
import openpyxl


class XLSXBuilder:
    def __init__(self, path=None):
        self.path = path or f'/tmp/{os.urandom(7).hex()}.xlsx'

    def create_file(self, data):
        data = set(data)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for row, item in enumerate(data, start=1):
            cell = worksheet.cell(row=row, column=1)
            cell.value = item

        workbook.save(self.path)
        workbook.close()

        return self.path

    def parse_file(self):
        workbook = openpyxl.load_workbook(self.path, data_only=True)
        worksheet = workbook.active
        rows = worksheet.rows
        values = [cell.value for row in rows for cell in row]
        return values

    def delete_file(self):
        if not os.path.exists(self.path):
            return

        os.remove(self.path)
